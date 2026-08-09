#!/usr/bin/env python3
"""Plain stdlib test runner: python3 tests/test_report.py — exits non-zero on failure."""

import base64
import hashlib
import json
import os
import re
import shutil
import sys
import tempfile
import time
import zlib
from datetime import datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

TMP = Path(tempfile.mkdtemp(prefix="countkit-rep-")).resolve()
os.environ["COUNTKIT_ROOT"] = str(TMP)

import app                                       # noqa: E402
import calib                                     # noqa: E402
import engine                                    # noqa: E402
import report                                    # noqa: E402
from fastapi.testclient import TestClient        # noqa: E402
from openpyxl import load_workbook               # noqa: E402

C = TestClient(app.app)
FAILS = []
SITE, DATE, EMPTY = "gerlache", "2026-08-04", "2026-08-05"
CAT = timezone(timedelta(hours=2))
GE, CH = "Great East Rd", "Church Rd"
STEM = f"countkit-{SITE}-{DATE}"


def check(name, fn):
    try:
        fn()
        print(f"ok   {name}")
    except Exception as e:
        FAILS.append(name)
        print(f"FAIL {name}: {e}")


def ts(hhmm, sec=0):
    """A local Africa/Lusaka wallclock on the survey date, as an epoch."""
    h, m = map(int, hhmm.split(":"))
    return datetime(2026, 8, 4, h, m, sec, tzinfo=CAT).timestamp()


def write(events):
    db = engine.connect(app.data_root(), SITE)
    db.execute("DELETE FROM events WHERE site=? AND date=?", (SITE, DATE))
    for cam, oid, cls, line, kind, t in events:
        db.execute("INSERT INTO events VALUES (?,?,?,?,?,?,?,?,?)",
                   (SITE, DATE, cam, oid, cls, line, kind, t, None))
    db.commit()
    db.close()


def segments(cam, epochs):
    """Segment names are written back through the same host-timezone strftime that
    engine.segment_epoch parses, so coverage lines up on a host in any timezone."""
    d = app.ingest_root() / DATE / SITE / cam
    d.mkdir(parents=True, exist_ok=True)
    for p in d.glob("*.mkv"):
        p.unlink()
    for e in epochs:
        (d / datetime.fromtimestamp(e).strftime("%Y%m%d-%H%M%S.mkv")).write_bytes(b"")


def calibrate():
    """cam1 owns Great East Rd (north arm), cam2 owns Church Rd (east arm)."""
    line = lambda kind, name, comp: {"kind": kind, "name": name, "compass": comp,
                                     "dir": "AB", "points": [[0, 100], [200, 100]]}
    calib.save_calibration(SITE, "cam1", {"image_size": [640, 480],
                                          "lines": [line("entry", GE, "N"),
                                                    line("exit", GE, "N")]})
    calib.save_calibration(SITE, "cam2", {"image_size": [640, 480],
                                          "lines": [line("entry", CH, "E"),
                                                    line("exit", CH, "E")]})


def seed():
    """One survey morning: 40/30/20/10 entries in the peak hour (hand-computable PHF),
    two buses for the heavy share, one cross-camera pair, then an hour with no footage."""
    evs, oid = [], 0
    for hhmm, n in (("07:00", 40), ("07:15", 30), ("07:30", 20), ("07:45", 10)):
        for k in range(n):
            oid += 1
            cls = "bus" if hhmm == "07:00" and k < 2 else "car"
            evs.append(("cam1", oid, cls, GE, "entry", ts(hhmm, k)))
    # The 07:00:39 car leaves through cam2's view of Church Rd 20 s later: no shared
    # tracker id, so pairing can only infer it — tier 2.
    evs.append(("cam2", 900, "car", CH, "exit", ts("07:00", 59)))
    write(evs)
    # Both cameras record 07:00–08:00, stop for an hour, come back at 09:00.
    segs = [ts(f"07:{m:02d}") for m in range(0, 60, 10)] + [ts("09:00"), ts("09:10")]
    segments("cam1", segs)
    segments("cam2", segs)


def build():
    r = C.post(f"/api/report/{SITE}/{DATE}")
    assert r.status_code == 200, (r.status_code, r.text)
    return r.json()


def bundle(name):
    return report.bundle_dir(app.data_root(), SITE, DATE) / name


def staging():
    """Per-build staging dirs, so one build never hashes another's bytes."""
    out = report.bundle_dir(app.data_root(), SITE, DATE)
    return sorted(p.name for p in out.parent.glob(out.name + ".tmp*"))


def pdf_streams(raw):
    """The drawn text, decoded. ReportLab writes page content through
    /ASCII85Decode + /FlateDecode, so nothing is greppable in the raw file."""
    out = []
    # Cut on the declared /Length, never on the next b"endstream": the ASCII85 alphabet
    # runs '!'..'u', so a payload can contain that word itself.
    for m in re.finditer(rb"/Length\s+(\d+)\s*>>\s*stream\r?\n", raw):
        body = raw[m.end():m.end() + int(m.group(1))]
        try:
            out.append(zlib.decompress(base64.a85decode(body.strip(), adobe=True)))
        except (ValueError, zlib.error):
            out.append(body)
    # WinAnsi bytes above 127 are written as octal escapes: \251 is the (c) glyph.
    text = re.sub(rb"\\(\d{3})", lambda m: bytes([int(m.group(1), 8)]), b"\n".join(out))
    # Parentheses delimit PDF string literals, so any in the copy arrive backslashed.
    return text.replace(rb"\(", b"(").replace(rb"\)", b")")


def cell_rows(ws):
    return [[c.value for c in row] for row in ws.iter_rows()]


# ---- the bundle ----

def export_writes_a_bundle():
    r = build()
    assert r["ok"] and r["site"] == SITE and r["date"] == DATE, r
    assert set(r["files"]) == {f"{STEM}.xlsx", f"{STEM}.pdf"}, r["files"]
    assert Path(r["dir"]) == report.bundle_dir(app.data_root(), SITE, DATE), r["dir"]
    assert not staging(), f"staging directories were left behind: {staging()}"


def manifest_hashes_verify():
    man = json.loads(bundle("manifest.json").read_text())
    for name, meta in man["files"].items():
        raw = bundle(name).read_bytes()
        assert hashlib.sha256(raw).hexdigest() == meta["sha256"], f"{name} hash mismatch"
        assert len(raw) == meta["bytes"], f"{name} size mismatch"
    assert man["qa"]["pairing_rate"] == 1.0, man["qa"]      # 1 of 100 entries paired
    assert man["qa"]["inferred_share"] == 100.0, man["qa"]
    assert man["generated_iso"].startswith("20"), man


def reexport_is_atomic_and_restamped():
    first = json.loads(bundle("manifest.json").read_text())
    time.sleep(1.05)                    # generated_iso has one-second resolution
    build()
    second = json.loads(bundle("manifest.json").read_text())
    assert second["generated_iso"] > first["generated_iso"], (first, second)
    assert set(second["files"]) == set(first["files"]), "re-export changed the file set"
    assert not staging(), staging()
    # A live rival build's staging dir must survive; only a crashed one is swept.
    out = report.bundle_dir(app.data_root(), SITE, DATE)
    rival = out.parent / (out.name + ".tmp-rival")
    rival.mkdir()
    (rival / "half-written.xlsx").write_bytes(b"PK")
    build()
    assert rival.exists(), "a concurrent build's staging directory was deleted"
    old = out.parent / (out.name + ".tmp-crashed")
    old.mkdir()
    os.utime(old, (time.time() - 7200, time.time() - 7200))
    build()
    assert not old.exists(), "a two-hour-old staging directory was not swept"
    shutil.rmtree(rival)


# ---- Excel ----

def excel_opens_with_the_four_sheets():
    wb = load_workbook(bundle(f"{STEM}.xlsx"))
    assert wb.sheetnames == ["Method", "Bins", "Movements", "Summary"], wb.sheetnames
    method = {r[0]: r[1] for r in cell_rows(wb["Method"]) if r and r[0]}
    assert method["Site"] == SITE and method["Date"] == DATE, method
    assert "two-tier" in method["Method"] and "inferred share" in method["Method"], method
    assert method["Cross-camera inferred %"] == 1.0, method
    assert method["Inferred share of paired %"] == 100.0, method
    assert method["QA flags"] == 0, method
    assert any(str(k).startswith("Coverage gap cam1") for k in method), method
    assert any(str(k).startswith("Class map") for k in method), method
    assert any(str(k).startswith("PCU factor") for k in method), method


def a_gap_bin_says_no_footage():
    rows = cell_rows(load_workbook(bundle(f"{STEM}.xlsx"))["Bins"])
    by = {r[0]: r for r in rows[2:]}
    assert "08:15" in by, sorted(str(k) for k in by)      # inside the 08:00–09:00 gap
    gap = by["08:15"][1:]
    assert gap and all(v == "no footage" for v in gap), gap
    assert 0 not in gap and None not in gap, "a gap must never read as a measured zero"
    # The same columns carry real numbers where there was footage.
    assert any(isinstance(v, int) and v for v in by["07:00"][1:]), by["07:00"]


def phf_matches_the_hand_computation():
    rows = cell_rows(load_workbook(bundle(f"{STEM}.xlsx"))["Bins"])
    hand = round(100 / (4 * 40), 2)                       # 40+30+20+10, busiest 40
    peak = next(r for r in rows if r[0] == "AM peak")
    phf = next(r for r in rows if r[0] == "PHF AM peak")
    assert 100 in peak[1:], peak
    assert hand in phf[1:], (phf, hand)
    summary = {r[0]: r[1] for r in cell_rows(load_workbook(bundle(f"{STEM}.xlsx"))["Summary"])}
    # The AM peak hour is wholly inside recorded footage, so its figures stand bare.
    assert summary["AM peak entering volume"] == 100, summary
    assert summary["AM peak PHF"] == hand, summary
    # Anything summed over the whole day crosses the 08:00–09:00 hole and must say so.
    assert summary["% heavy (bus + heavy truck)"] == "2.0*", summary   # 2 buses in 100
    assert summary[f"Entering {GE}"] == "100*", summary
    assert summary["Day total entering"] == "100*", summary


def a_day_total_over_a_gap_is_marked():
    """A headline number that silently swallows the 08:00–09:00 hole contradicts the
    same workbook's 'no footage' cells — every such figure carries a star and a note."""
    wb = load_workbook(bundle(f"{STEM}.xlsx"))
    for sheet in ("Bins", "Summary"):
        flat = [str(v) for r in cell_rows(wb[sheet]) for v in r if v is not None]
        note = [v for v in flat if v.startswith("* excludes")]
        assert len(note) == 1, (sheet, note)
        assert re.fullmatch(r"\* excludes \d+ no-footage bins — see coverage", note[0]), note
        assert any(v.endswith("*") and v != note[0] for v in flat), (sheet, "nothing marked")
    day = next(r for r in cell_rows(wb["Bins"]) if r[0] == "Day total")
    assert "100*" in [str(v) for v in day[1:]], day
    # The AM peak hour is wholly inside recorded footage: no star, a plain number.
    peak = next(r for r in cell_rows(wb["Bins"]) if r[0] == "AM peak")
    assert 100 in peak[1:], peak


def a_missing_peak_is_named_not_zeroed():
    """The fixture has no afternoon traffic. An all-zero PM matrix would claim nobody
    turned; the truth is there was no PM peak hour at all."""
    flat = [str(v) for r in cell_rows(load_workbook(bundle(f"{STEM}.xlsx"))["Movements"])
            for v in r if v is not None]
    assert "no PM peak recorded" in flat, flat
    assert not any(v.startswith("PM peak — entry") for v in flat), flat
    assert any(v.startswith("AM peak — entry") for v in flat), "the AM peak still renders"


def a_formula_arm_name_cannot_execute_or_crash():
    """Arm names are operator input. openpyxl runs a leading '=' as a live formula and
    refuses control characters outright, killing the bundle over one stray byte."""
    assert report._cell("=1+1") == "'=1+1"
    for bad in ("+SUM(A1)", "-2", "@cmd"):
        assert report._cell(bad).startswith("'"), bad
    assert report._cell("North\x01 Rd\x1f") == "North Rd", report._cell("North\x01 Rd\x1f")
    assert report._cell("Great East Rd") == "Great East Rd"
    assert report._cell(7) == 7, "numbers pass through untouched"


def an_inferred_cell_is_labelled():
    rows = cell_rows(load_workbook(bundle(f"{STEM}.xlsx"))["Movements"])
    flat = [v for r in rows for v in r]
    assert "1 (1 inf)" in flat, [v for v in flat if v]
    assert any(str(v).startswith("AM peak — entry") for v in flat), flat[:8]
    assert any(str(v).startswith("Full day — entry") for v in flat), "day block missing"


# ---- PDF ----

def pdf_is_a_real_pdf():
    raw = bundle(f"{STEM}.pdf").read_bytes()
    assert raw.startswith(b"%PDF"), raw[:8]
    text = pdf_streams(raw)
    # A page of compressed vector drawing is ~3 KB on the wire; measure the drawing
    # itself rather than the compressed size, which a stub PDF could also reach.
    assert len(raw) > 2048, f"{len(raw)} bytes on disk"
    assert len(text) > 3000, f"{len(text)} bytes of content — the page is nearly empty"
    # reportlab writes each stroked segment inline as "x y m x y l S". Match the whole
    # operator, not a bare "l", which drawn text could also supply. This fixture is a
    # two-arm site: 8 road edges + 6 turn arrows (22 segments) + the north arrow.
    assert text.count(b" l S") >= 25, f"{text.count(b' l S')} lines — no junction diagram"
    assert b"Dexterity" in text, "branding company_name is missing from the header"
    assert b"arrows kerb-first (LHT)" in text, "the arrow-order note is missing"
    assert b"AM peak movement matrix" in text, "the movement matrix is missing"
    assert b"largest coverage gap: cam1" in text, "the gap footnote names no camera"


def probe_page_only_when_configured():
    plain = pdf_streams(bundle(f"{STEM}.pdf").read_bytes())
    assert b"\xa9 tomtom" not in plain and b"Probe corroboration" not in plain, (
        "probe attribution appeared with no dataset configured")

    csv_path = TMP / "probe.csv"
    csv_path.write_text(
        "site,arm,bin_start_iso,delay_s,speed_kmh,sample_n\n"
        + "".join(f"{SITE},{GE},2026-08-04T07:{m:02d}:00+02:00,{40 + m},18.0,{12 + m}\n"
                  for m in (0, 15, 30, 45)))
    cfg = {**app.CONFIG, "probe": {"provider": "tomtom", "dataset": str(csv_path)}}
    report.build(SITE, DATE, app.data_root(), app.ingest_root(), cfg)
    text = pdf_streams(bundle(f"{STEM}.pdf").read_bytes())
    assert b"Probe corroboration" in text, "no corroboration page with a dataset configured"
    assert b"\xa9 tomtom" in text, "provider attribution missing"
    assert b"licensed floating-car data" in text and b"not a second count" in text, (
        "the licence line must say what the probe data is not")
    assert b"median probe sample n=" in text, "per-arm sample size missing"
    report.build(SITE, DATE, app.data_root(), app.ingest_root(), app.CONFIG)   # back to plain


# ---- routes ----

def files_come_only_from_the_manifest():
    r = C.get(f"/api/report/file/{SITE}/{DATE}/{STEM}.xlsx")
    assert r.status_code == 200 and r.content[:2] == b"PK", r.status_code
    assert C.get(f"/api/report/file/{SITE}/{DATE}/{STEM}.pdf").content[:4] == b"%PDF"
    # The manifest vouches for the two deliverables and nothing else — not even itself.
    assert C.get(f"/api/report/file/{SITE}/{DATE}/manifest.json").status_code == 404
    assert C.get(f"/api/report/file/{SITE}/{DATE}/nope.xlsx").status_code == 404
    for bad in ("..%2f..%2fconfig.yaml", "..%2F..%2F..%2Fetc%2Fpasswd", "%2e%2e%2fapp.py"):
        r = C.get(f"/api/report/file/{SITE}/{DATE}/{bad}")
        assert r.status_code != 200, (bad, r.status_code)
        assert b"ingest_root" not in r.content and b"FastAPI" not in r.content, bad


def status_reads_the_manifest():
    s = C.get(f"/api/report/{SITE}/{DATE}").json()
    assert s["exists"] and set(s["files"]) == {f"{STEM}.xlsx", f"{STEM}.pdf"}, s
    none = C.get(f"/api/report/{SITE}/{EMPTY}").json()
    assert none["exists"] is False and none["files"] == {}, none


def nothing_analyzed_refuses():
    r = C.post(f"/api/report/{SITE}/{EMPTY}")
    assert r.status_code == 400, r.status_code
    assert "nothing analyzed" in r.json()["detail"], r.json()
    assert not report.bundle_dir(app.data_root(), SITE, EMPTY).exists(), (
        "a refused export must leave no half-written bundle")


# ---- turn geometry ----

def turns_read_left_hand_traffic():
    # Entering from the north arm you are heading south: the east arm is a left turn,
    # the south arm straight on, the west arm a right turn across oncoming traffic.
    assert report.turn_of("N", "E") == "left"
    assert report.turn_of("N", "S") == "straight"
    assert report.turn_of("N", "W") == "right"
    assert report.turn_of("N", "N") == "uturn"
    assert report.turn_of("W", "N") == "left" and report.turn_of("W", "S") == "right"
    assert report.turn_of("", "N") is None, "an untagged arm cannot be given a turn"


# ---- the Report tab ----

def report_tab_offers_the_export():
    html = C.get("/").text
    sec = html.split('<section id="report"', 1)[1].split("</section>", 1)[0]
    for bit in ('id="r-day"', 'id="r-go"', 'id="rcard"'):
        assert bit in sec, bit
    assert "/api/report/file/" in html, "no download links to the bundle"
    assert "#/analyze" in html, "an unexportable day must point at Analyze"
    assert "SHA256 MANIFEST" in html, "the custody line is missing from the done state"


calibrate()
seed()
check("export writes a bundle of exactly the two deliverables", export_writes_a_bundle)
check("manifest sha256s and sizes verify", manifest_hashes_verify)
check("re-export restamps and swaps in one move", reexport_is_atomic_and_restamped)
check("workbook opens with Method, Bins, Movements, Summary", excel_opens_with_the_four_sheets)
check("a bin with no footage says so, never zero", a_gap_bin_says_no_footage)
check("peak hour and PHF match the hand computation", phf_matches_the_hand_computation)
check("a total spanning a gap is starred and footnoted", a_day_total_over_a_gap_is_marked)
check("a half-day with no peak says so, never a zero grid", a_missing_peak_is_named_not_zeroed)
check("a formula-shaped arm name is defused", a_formula_arm_name_cannot_execute_or_crash)
check("an inferred movement is labelled in its cell", an_inferred_cell_is_labelled)
check("PDF renders with the diagram, matrix and footnotes", pdf_is_a_real_pdf)
check("probe attribution appears only with a dataset", probe_page_only_when_configured)
check("downloads are limited to manifested names", files_come_only_from_the_manifest)
check("status reads the manifest; absent is absent", status_reads_the_manifest)
check("an unanalyzed site-day refuses to export", nothing_analyzed_refuses)
check("turn classification reads left-hand traffic", turns_read_left_hand_traffic)
check("Report tab carries the export controls", report_tab_offers_the_export)

print(f"\n{'FAILED: ' + ', '.join(FAILS) if FAILS else 'all passed'}")
sys.exit(1 if FAILS else 0)
