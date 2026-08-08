#!/usr/bin/env python3
"""The client deliverable: Excel workbook + PDF, hashed and manifested like the footage.

Every number here traces back to the events table; nothing is editable by hand."""

import hashlib
import json
import shutil
import textwrap
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdfcanvas

import aggregate
import calib

COMPASS = {"N": 0, "E": 90, "S": 180, "W": 270}
HEAVY = ("heavy_truck", "bus")
METHOD = ("Counts from camera analysis (CountKit); movements paired two-tier: "
          "same-camera tracker + cross-camera class/transit-time inference — "
          "inferred share reported per cell.")
LICENCE = ("licensed floating-car data — corroboration of survey timing and peaks, "
           "not a second count")
SCHEMATIC = ("arms without compass tags are placed in schematic order "
             "(arm name, clockwise from north) — turns are labelled by that order")


def turn_of(entry_compass, exit_compass):
    """Which of the three arrows a movement belongs to, for left-hand traffic.

    An arm's compass tag says where the arm lies, and a vehicle entering from it is
    already heading the other way — but the tag cancels out of the difference, so the
    approach heading is the ENTRY arm's tag as written (adding 180 to it would shift
    every turn by a half-turn and swap left with right). Clockwise difference:
    90 left, 180 straight, 270 right, 0 U-turn."""
    a, b = COMPASS.get(entry_compass or ""), COMPASS.get(exit_compass or "")
    if a is None or b is None:
        return None
    return {90: "left", 180: "straight", 270: "right", 0: "uturn"}[(b - a) % 360]


def _bins_in(d, period):
    p = d["peaks"].get(period)
    if not p:
        return []
    i = next((k for k, b in enumerate(d["bins"]) if b["start"] == p["start"]), -1)
    return [b["ts"] for b in d["bins"][i:i + 4]] if i >= 0 else []


def _cells(d, period):
    keep = None if period == "day" else _bins_in(d, period)
    return [c for c in d["movements"]["od"] if keep is None or c["bin"] in keep]


def _phf(d, arm, bins):
    rows = [b for b in d["bins"] if b["ts"] in bins]
    v = [(b["arms"][arm]["total"] if arm else b["total"]) for b in rows]
    return round(sum(v) / (4 * max(v)), 2) if v and max(v) else None


def _heavy_pct(d):
    tot = heavy = 0
    for b in d["bins"]:
        for arm in b["arms"].values():
            for c, n in arm["classes"].items():
                tot += n
                heavy += n if c in HEAVY else 0
    return round(heavy / tot * 100, 1) if tot else 0.0


# ---------------------------------------------------------------- Excel

def _excel(path, d, config):
    wb = Workbook()
    bold = Font(bold=True)
    arms, classes = d["arms"], d["classes"]
    qa, pr = d["qa"], d["qa"]["pairing"]

    ws = wb.active
    ws.title = "Method"
    rows = [("Site", d["site"]), ("Date", d["date"]),
            ("Company", (config.get("branding") or {}).get("company_name", "")),
            ("Cameras", ", ".join(qa["coverage"]["per_cam"])),
            ("Survey window", " to ".join(qa["coverage"]["window_iso"])),
            ("Coverage hours expected", qa["coverage"]["expected_hours"]),
            ("Pairing rate %", pr["rate"]), ("Same-camera %", pr["same"]),
            ("Cross-camera inferred %", pr["inferred"]),
            ("Inferred share of paired %", pr["inferred_share"]),
            ("QA flags", pr["flagged"]), ("Method", METHOD)]
    for cam, off in qa["offsets"].items():
        rows.append((f"Clock offset {cam} (s)", off))
    for cam, c in qa["coverage"]["per_cam"].items():
        rows.append((f"Recorded hours {cam}", c["hours"]))
        for g in c["gaps_iso"]:
            rows.append((f"Coverage gap {cam}", " to ".join(g)))
    for c in (config.get("classes") or []):
        rows.append((f"Class map {c.get('model')}", c.get("report")))
    for k, v in (config.get("pcu") or {}).items():
        rows.append((f"PCU factor {k}", v))
    if qa["probe"]:
        rows += [("Probe provider", qa["probe"]["provider"]), ("Probe licence", LICENCE)]
    for r in rows:
        ws.append(list(r))
    for c in ws["A"]:
        c.font = bold
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    ws = wb.create_sheet("Bins")
    head1, head2 = ["Bin"], [""]
    for a in arms:
        head1 += [a] + [""] * len(classes)
        head2 += classes + ["total"]
    ws.append(head1)
    ws.append(head2)
    for c in ws[1] + ws[2]:
        c.font = bold
    for b in d["bins"]:
        row = [b["start"][11:16]]
        for a in arms:
            cell = b["arms"][a]
            if cell["gap"]:
                # Never 0, never blank: a gap is missing footage, not measured silence.
                row += ["no footage"] * (len(classes) + 1)
            else:
                row += [cell["classes"].get(c, 0) for c in classes] + [cell["total"]]
        ws.append(row)
    for label, period in (("AM peak", "am"), ("PM peak", "pm")):
        bins = _bins_in(d, period)
        if not bins:
            continue
        rows_ = [b for b in d["bins"] if b["ts"] in bins]
        tot, phf = [label], ["PHF " + label]
        for a in arms:
            tot += [""] * len(classes) + [sum(b["arms"][a]["total"] for b in rows_)]
            phf += [""] * len(classes) + [_phf(d, a, bins)]
        ws.append(tot)
        ws.append(phf)
    day = ["Day total"]
    for a in arms:
        day += [""] * len(classes) + [sum(b["arms"][a]["total"] for b in d["bins"])]
    ws.append(day)

    ws = wb.create_sheet("Movements")
    for label, period in (("AM peak", "am"), ("PM peak", "pm"), ("Full day", "day")):
        cells = _cells(d, period)
        ws.append([f"{label} — entry (rows) x exit (cols)"])
        ws.cell(ws.max_row, 1).font = bold
        ws.append(["entry \\ exit"] + arms)
        grid = {}
        for c in cells:
            g = grid.setdefault((c["entry"], c["exit"]), [0, 0])
            g[0] += c["count"]
            g[1] += c["tier2_count"]
        for en in arms:
            row = [en]
            for ex in arms:
                n, inf = grid.get((en, ex), (0, 0))
                row.append(f"{n} ({inf} inf)" if inf else n)
            ws.append(row)
        ws.append([])

    ws = wb.create_sheet("Summary")
    ws.append(["Metric", "Value"])
    ws.cell(1, 1).font = bold
    ws.cell(1, 2).font = bold
    for label, period in (("AM", "am"), ("PM", "pm")):
        p = d["peaks"].get(period)
        ws.append([f"{label} peak start", p["start"] if p else "none"])
        ws.append([f"{label} peak entering volume", p["total"] if p else 0])
        ws.append([f"{label} peak PHF", p["phf"] if p else ""])
    ws.append(["Day total entering", sum(b["total"] for b in d["bins"])])
    ws.append(["% heavy (bus + heavy truck)", _heavy_pct(d)])
    for a in arms:
        ws.append([f"Entering {a}", sum(b["arms"][a]["total"] for b in d["bins"])])
        ws.append([f"AM PHF {a}", _phf(d, a, _bins_in(d, "am")) or ""])
    wb.save(path)


# ---------------------------------------------------------------- PDF

def _arrow(c, pts):
    """Polyline with a head on the last segment — a turn is one bent arrow, so the
    reader sees where the movement goes without decoding the label."""
    for a, b in zip(pts, pts[1:]):
        c.line(a[0], a[1], b[0], b[1])
    (x0, y0), (x1, y1) = pts[-2], pts[-1]
    n = ((x1 - x0) ** 2 + (y1 - y0) ** 2) ** .5 or 1
    ux, uy = (x1 - x0) / n, (y1 - y0) / n
    px, py = -uy, ux
    c.line(x1, y1, x1 - ux * 5 + px * 2.6, y1 - uy * 5 + py * 2.6)
    c.line(x1, y1, x1 - ux * 5 - px * 2.6, y1 - uy * 5 - py * 2.6)


def _diagram(c, d, x0, y0, size):
    """Plan view, north up, three arrows per approach drawn kerb-first for left-hand
    traffic: the left hook sits nearest the kerb, then straight, then the right hook."""
    cx, cy = x0 + size / 2, y0 + size / 2
    box = size * .18
    c.rect(cx - box, cy - box, box * 2, box * 2)
    c.setFont("Helvetica", 8)
    c.drawString(x0, y0 + size - 6, "N")
    c.line(x0 + 4, y0 + size - 16, x0 + 4, y0 + size - 8)
    c.line(x0 + 4, y0 + size - 8, x0 + 1.5, y0 + size - 12)
    c.line(x0 + 4, y0 + size - 8, x0 + 6.5, y0 + size - 12)

    arms = d["arms"]
    tags = {a: (d.get("compass") or {}).get(a, "") for a in arms}
    placed, free = {}, [t for t in ("N", "E", "S", "W")]
    for a in arms:                       # compass tags win; the rest fill in clockwise
        t = tags.get(a)
        if t in free:
            placed[a] = t
            free.remove(t)
    for a in arms:
        if a not in placed and free:
            placed[a] = free.pop(0)

    dirs = {"N": (0, 1), "E": (1, 0), "S": (0, -1), "W": (-1, 0)}
    # Hooks stop at the far kerb rather than sailing past it, which is what made the
    # three approaches read as one tangle.
    lane, stop, out = 16, box * .95, box * 1.25 + 22
    # Road edges for every arm position, so the plan reads as a junction rather than
    # as arrows floating around a square.
    for ux, uy in dirs.values():
        px, py = -uy, ux
        for side in (1, -1):
            c.line(cx + ux * box + px * box * side, cy + uy * box + py * box * side,
                   cx + ux * (size / 2) + px * box * side,
                   cy + uy * (size / 2) + py * box * side)
    cells = _cells(d, "am")
    for arm, tag in placed.items():
        ux, uy = dirs[tag]
        # The vehicle drives inward (-u), so +p is its right hand and -p the near kerb.
        px, py = -uy, ux
        turns = {"left": [0, 0], "straight": [0, 0], "right": [0, 0]}
        for cell in cells:
            if cell["entry"] != arm:
                continue
            t = turn_of(tags.get(arm), tags.get(cell["exit"]))
            if t in turns:                       # U-turns carry no arrow
                turns[t][0] += cell["count"]
                turns[t][1] += cell["count"] if cell["cls"] in HEAVY else 0
        # Kerb-first for left-hand traffic: the left hook runs nearest the kerb, then
        # the straight-ahead lane, then the right hook crossing oncoming traffic.
        for i, t in enumerate(("left", "straight", "right")):
            n, heavy = turns[t]
            off = (i - 1) * lane
            start = (cx + ux * out + px * off, cy + uy * out + py * off)
            if t == "straight":
                pts = [start, (cx - ux * box * .7 + px * off, cy - uy * box * .7 + py * off)]
            else:
                dx, dy = (-px, -py) if t == "left" else (px, py)
                deep = box * .75 if t == "left" else -box * .35
                bend = (cx + ux * deep + px * off, cy + uy * deep + py * off)
                pts = [start, bend, (bend[0] + dx * stop, bend[1] + dy * stop)]
            _arrow(c, pts)
            c.setFont("Helvetica", 6.5)
            # Each count sits at the tail of its own arrow, in lane order across the
            # approach, so the reader never has to guess which arrow a number belongs to.
            c.drawCentredString(cx + ux * (out + 8) + px * off,
                                cy + uy * (out + 8) + py * off - 2,
                                f"{n} ({heavy})" if heavy else str(n))
        c.setFont("Helvetica", 7)
        # Beside the number row, not above it: on an east/west approach a centred name
        # sits at the same height as its own counts and prints over them.
        c.drawCentredString(cx + ux * (out + 9) + px * lane * 2.6,
                            cy + uy * (out + 9) + py * lane * 2.6 - 2,
                            f"{arm[:14]} ({tag})")


def _pdf(path, d, config):
    brand = config.get("branding") or {}
    qa, pr = d["qa"], d["qa"]["pairing"]
    # Uncompressed: the deliverable stays greppable and diffable, so an auditor can
    # pull the text out years later without the tooling we happened to ship with.
    c = pdfcanvas.Canvas(str(path), pagesize=A4, pageCompression=0)
    W, H = A4
    y = H - 20 * mm
    c.setFont("Helvetica-Bold", 13)
    c.drawString(20 * mm, y, f"{brand.get('company_name', 'CountKit')} · {d['site']}"
                             f" · {d['date']}")
    logo = brand.get("logo_path") or ""
    if logo and Path(logo).is_file():
        try:
            c.drawImage(logo, W - 45 * mm, y - 4 * mm, width=25 * mm, height=10 * mm,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass          # a bad logo must never cost the client their report
    y -= 7 * mm
    c.setFont("Helvetica", 8)
    c.drawString(20 * mm, y, "Junction turning-movement survey · counts from camera "
                             "analysis · Africa/Lusaka")

    _diagram(c, d, 20 * mm, y - 105 * mm, 100 * mm)

    bx, by = 130 * mm, y - 30 * mm
    peak = d["peaks"].get("am")
    c.setFont("Helvetica-Bold", 8)
    c.drawString(bx, by, "AM PEAK")
    c.setFont("Helvetica", 8)
    for i, line in enumerate([
            f"interval {peak['start'][11:16] if peak else '—'}",
            f"total entering {peak['total'] if peak else 0}",
            f"PHF {peak['phf'] if peak else '—'}",
            f"% heavy {_heavy_pct(d)}",
            f"pairing {pr['rate']}% ({pr['inferred']}% inferred)",
            "arrows kerb-first (LHT)"]):
        c.drawString(bx, by - 5 * mm - i * 4.4 * mm, line)
    c.rect(bx - 3 * mm, by - 32 * mm, 55 * mm, 38 * mm)

    # Headline numerals: the four figures a reader should get without reading anything.
    hy = by - 46 * mm
    for i, (num, cap) in enumerate([
            (str(peak["total"] if peak else 0), "AM PEAK ENTERING"),
            (str(peak["phf"] if peak else "—"), "PHF"),
            (str(sum(b["total"] for b in d["bins"])), "DAY TOTAL"),
            (f"{_heavy_pct(d)}%", "HEAVY")]):
        ry = hy - i * 14 * mm
        c.setFont("Helvetica-Bold", 17)
        c.drawString(bx, ry, num)
        c.setFont("Helvetica", 6)
        c.drawString(bx, ry - 4 * mm, cap)

    ty = y - 118 * mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(20 * mm, ty, "AM peak movement matrix")
    ty -= 6 * mm
    grid = {}
    for cell in _cells(d, "am"):
        g = grid.setdefault((cell["entry"], cell["exit"]), [0, 0])
        g[0] += cell["count"]
        g[1] += cell["tier2_count"]
    arms = d["arms"]
    c.setFont("Helvetica", 7)
    colw = min(28 * mm, (W - 60 * mm) / max(1, len(arms)))
    c.drawString(20 * mm, ty, "entry \\ exit")
    for j, ex in enumerate(arms):
        c.drawString(52 * mm + j * colw, ty, ex[:16])
    for i, en in enumerate(arms):
        ry = ty - (i + 1) * 5 * mm
        c.drawString(20 * mm, ry, en[:18])
        for j, ex in enumerate(arms):
            n, inf = grid.get((en, ex), (0, 0))
            c.drawString(52 * mm + j * colw, ry,
                         f"{n} ({inf} inf)" if inf else str(n))

    fy = 24 * mm
    c.setFont("Helvetica", 6.5)
    notes = [METHOD]
    if not all((d.get("compass") or {}).get(a) for a in arms):
        notes.append(SCHEMATIC)
    # Largest by duration, not the last one of the day — the longest blind spell is the
    # one a reader has to know about.
    worst = max(((cam, g, iso) for cam, cv in qa["coverage"]["per_cam"].items()
                 for g, iso in zip(cv["gaps"], cv["gaps_iso"])),
                key=lambda x: x[1][1] - x[1][0], default=None)
    if worst:
        mins = round((worst[1][1] - worst[1][0]) / 60)
        notes.append(f"largest coverage gap: {worst[0]} {worst[2][0]} to {worst[2][1]}"
                     f" ({mins} min of no footage)")
    weak = [cam for cam, c in pr["per_cam"].items() if c["rate"] < aggregate.OK_RATE]
    if pr["state"] != "ok" or weak:
        notes.append(f"pairing rate {pr['rate']}% ({pr['inferred_share']}% of pairs "
                     f"inferred) · {pr['flagged']} movements flagged in QA"
                     + (f" · below the {aggregate.OK_RATE}% threshold on "
                        + ", ".join(weak) if weak else ""))
    # Wrapped, never clipped: a method statement cut mid-word reads as a broken report.
    line = 0
    for n in notes:
        for part in textwrap.wrap(n, 150) or [""]:
            c.drawString(20 * mm, fy - line * 3.4 * mm, part)
            line += 1

    if qa["probe"]:
        c.showPage()
        _probe_page(c, d, W, H)
    c.showPage()
    c.save()


def _probe_page(c, d, W, H):
    """Corroboration only — the provider series never merges into a counted one."""
    pr = d["qa"]["probe"]
    y = H - 20 * mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, y, "Probe corroboration")
    c.setFont("Helvetica", 8)
    c.drawString(20 * mm, y - 6 * mm, "Measured volumes (bars) against provider delay "
                                      "(dashed). Corroboration of timing and peaks only.")
    top = y - 20 * mm
    for arm in d["arms"]:
        # Probe rows are keyed by the provider's spelling of the arm, folded to lower
        # case at load; the report's arm names keep their display casing.
        cells = {x["bin"]: x for x in pr["cells"] if x["arm"] == arm.strip().lower()}
        vols = [b["arms"][arm]["total"] for b in d["bins"]]
        vmax = max(vols + [1])
        dmax = max([x["delay_s"] for x in cells.values()] + [1])
        c.setFont("Helvetica-Bold", 8)
        c.drawString(20 * mm, top, arm)
        ns = [x["sample_n"] for x in cells.values()]
        c.setFont("Helvetica", 7)
        c.drawString(70 * mm, top, f"median probe sample n={sorted(ns)[len(ns) // 2] if ns else 0}"
                                   f" · {len(ns)} of {len(d['bins'])} bins")
        w = (W - 50 * mm) / max(1, len(d["bins"]))
        for i, b in enumerate(d["bins"]):
            x = 20 * mm + i * w
            c.rect(x, top - 22 * mm, w * .7, vols[i] / vmax * 18 * mm, fill=0)
        c.setDash(3, 2)
        prev = None
        for i, b in enumerate(d["bins"]):
            cell = cells.get(b["ts"])
            if not cell:
                prev = None
                continue          # broken where the provider has no data
            x = 20 * mm + i * w + w * .35
            py = top - 22 * mm + cell["delay_s"] / dmax * 18 * mm
            if prev:
                c.line(prev[0], prev[1], x, py)
            prev = (x, py)
        c.setDash()
        top -= 34 * mm
        if top < 40 * mm:
            break
    c.setFont("Helvetica", 7)
    c.drawString(20 * mm, 22 * mm, f"© {pr['provider']}")
    c.drawString(20 * mm, 18 * mm, LICENCE)


# ---------------------------------------------------------------- bundle

def bundle_dir(data_root, site, date):
    return Path(data_root) / "reports" / site / date


def _sha(p):
    h = hashlib.sha256()
    h.update(p.read_bytes())
    return h.hexdigest()


def build(site, date, data_root, ingest_root, config):
    d = aggregate.counts(site, date, data_root, ingest_root, config)
    # Compass tags live on the calibrations, not in the counts payload.
    d["compass"] = {a["arm"]: a["compass"] for a in calib.arm_map(site)["arms"]}
    if not any(b["total"] for b in d["bins"]):
        raise ValueError(f"{site} {date}: nothing analyzed — run it in Analyze first")
    out = bundle_dir(data_root, site, date)
    tmp = out.with_name(out.name + ".tmp")
    shutil.rmtree(tmp, ignore_errors=True)
    tmp.mkdir(parents=True)
    stem = f"countkit-{site}-{date}"
    _excel(tmp / f"{stem}.xlsx", d, config)
    _pdf(tmp / f"{stem}.pdf", d, config)
    pr = d["qa"]["pairing"]
    man = {"generated_iso": datetime.now(aggregate.CAT).isoformat(timespec="seconds"),
           "site": site, "date": date,
           "qa": {"pairing_rate": pr["rate"], "inferred_share": pr["inferred_share"],
                  "coverage_hours": d["qa"]["coverage"]["expected_hours"],
                  "flagged": pr["flagged"]},
           "files": {p.name: {"sha256": _sha(p), "bytes": p.stat().st_size}
                     for p in sorted(tmp.iterdir())}}
    (tmp / "manifest.json").write_text(json.dumps(man, indent=2))
    # Same custody discipline as the footage: the bundle appears complete or not at all.
    shutil.rmtree(out, ignore_errors=True)
    tmp.rename(out)
    return {"ok": True, "dir": str(out), **man}


def status(data_root, site, date):
    p = bundle_dir(data_root, site, date) / "manifest.json"
    try:
        return {"exists": True, **json.loads(p.read_text())}
    except (OSError, ValueError):
        return {"exists": False, "files": {}}
