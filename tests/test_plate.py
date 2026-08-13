#!/usr/bin/env python3
"""Plain stdlib test runner: python3 tests/test_plate.py — exits non-zero on failure.

The plate-confirmed pairing tier: matched licence plates are evidence, bound only
when unambiguous, priced conservatively per cell, and never exported anywhere."""

import json
import os
import sys
import tempfile
from datetime import datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

TMP = Path(tempfile.mkdtemp(prefix="countkit-plate-")).resolve()
os.environ["COUNTKIT_ROOT"] = str(TMP)

import aggregate                                # noqa: E402
import app                                      # noqa: E402
import calib                                    # noqa: E402
import engine                                   # noqa: E402
import report                                   # noqa: E402
from fastapi.testclient import TestClient       # noqa: E402

C = TestClient(app.app)
FAILS = []
SITE, DATE = "gerlache", "2026-08-04"
CAT = timezone(timedelta(hours=2))
GE, CH = "Great East Rd", "Church Rd"


def check(name, fn):
    try:
        fn()
        print(f"ok   {name}")
    except Exception as e:
        FAILS.append(name)
        print(f"FAIL {name}: {e}")


def ts(hhmm, sec=0):
    h, m = map(int, hhmm.split(":"))
    return datetime(2026, 8, 4, h, m, sec, tzinfo=CAT).timestamp()


def write(events):
    db = engine.connect(app.data_root(), SITE)
    db.execute("DELETE FROM events WHERE site=? AND date=?", (SITE, DATE))
    for cam, oid, cls, line, kind, t in events:
        db.execute("INSERT INTO events VALUES (?,?,?,?,?,?,?,?,?)",
                   (SITE, DATE, cam, oid, cls, line, kind, t, None))
    db.commit()
    db.close()


def plates(rows):
    p = TMP / "data" / "plate" / f"{SITE}-{DATE}.jsonl"
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text("".join(json.dumps(r) + "\n" for r in rows))


def prow(cam, oid, line, kind, t, plate, conf=0.9):
    return {"cam": cam, "obj_id": oid, "line": line, "kind": kind, "ts": t,
            "plate": plate, "conf": conf}


def calibrate():
    line = lambda kind, name: {"kind": kind, "name": name, "compass": "", "dir": "AB",
                               "points": [[0, 100], [200, 100]]}
    calib.save_calibration(SITE, "cam1", {"image_size": [640, 480],
                                          "lines": [line("entry", GE), line("exit", GE)]})
    calib.save_calibration(SITE, "cam2", {"image_size": [640, 480],
                                          "lines": [line("entry", CH), line("exit", CH)]})


def segments(cam, hhmms):
    d = TMP / "ingest" / DATE / SITE / cam
    d.mkdir(parents=True, exist_ok=True)
    for hhmm in hhmms:
        h, m = hhmm.split(":")
        (d / f"20260804-{h}{m}00.mkv").write_bytes(b"")


def run():
    return aggregate.counts(SITE, DATE, app.data_root(), TMP / "ingest", app.CONFIG)


calibrate()
segments("cam1", ["08:00"])
segments("cam2", ["08:00"])


# Two same-class vehicles 2 s apart: the ambiguity gate refuses this contest, which
# is exactly where a plate must be able to settle it.
AMBIG = [("cam1", 1, "car", GE, "entry", ts("08:00")),
         ("cam1", 2, "car", GE, "entry", ts("08:00", 2)),
         ("cam2", 51, "car", CH, "exit", ts("08:00", 30)),
         ("cam2", 52, "car", CH, "exit", ts("08:00", 32))]


def a_plate_settles_what_the_gate_refused():
    write(AMBIG)
    plates([])
    d = run()
    assert d["movements"]["tier3"] == 0 and d["movements"]["tier2"] == 0, d["movements"]
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4021")])
    d = run()
    assert d["movements"]["tier3"] == 1, d["movements"]
    pr = d["qa"]["pairing"]
    assert pr["plate"] == {"reads": 2, "resolved": 2, "pairs": 1}, pr["plate"]
    # The cascade compounds: with the plate claiming one contestant, the OTHER
    # entry/exit no longer has a rival, so tier 2 may now pair it — one plate read
    # converted BOTH halves of an ambiguous contest. And the rate counts plate pairs
    # as paired: it rises, never falls.
    assert d["movements"]["tier2"] == 1, d["movements"]
    assert pr["paired"] == 2 and pr["inferred_share"] == 50.0, pr


def one_ocr_slip_is_tolerated_two_are_not():
    write(AMBIG)
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4027")])   # distance 1
    assert run()["movements"]["tier3"] == 1
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAF4027")])   # distance 2
    assert run()["movements"]["tier3"] == 0


def a_plate_in_reach_of_two_exits_binds_nothing():
    write(AMBIG)
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4021"),
            prow("cam2", 52, CH, "exit", ts("08:00", 32), "BAE4028")])   # both within 1
    d = run()
    assert d["movements"]["tier3"] == 0, d["movements"]
    # And the honest failure is visible: reads resolved, no pair claimed.
    assert d["qa"]["pairing"]["plate"]["resolved"] == 3


def a_low_confidence_read_is_no_read():
    write(AMBIG)
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021", conf=0.4),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4021")])
    d = run()
    assert d["movements"]["tier3"] == 0
    assert d["qa"]["pairing"]["plate"] == {"reads": 2, "resolved": 1, "pairs": 0}


def plates_outrank_a_one_sided_refinement():
    write(AMBIG)
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4021")])
    (TMP / "data" / "refine").mkdir(parents=True, exist_ok=True)
    (TMP / "data" / "refine" / f"{SITE}-{DATE}.jsonl").write_text(json.dumps(
        {"cam": "cam1", "obj_id": 1, "line": GE, "kind": "entry", "ts": ts("08:00"),
         "to": "lgv"}) + "\n")
    d = run()
    assert d["movements"]["tier3"] == 1, "a class edit unpaired hard plate evidence"
    (TMP / "data" / "refine" / f"{SITE}-{DATE}.jsonl").write_text("")


def a_stale_sidecar_shows_reads_without_resolution():
    # Re-analysis renumbered every tracker id: rows resolve to nothing, pair nothing,
    # and the QA surface says so instead of the tier silently vanishing.
    write([("cam1", 901, "car", GE, "entry", ts("08:00")),
           ("cam2", 902, "car", CH, "exit", ts("08:00", 30))])
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4021")])
    d = run()
    assert d["movements"]["tier3"] == 0
    assert d["qa"]["pairing"]["plate"] == {"reads": 2, "resolved": 0, "pairs": 0}


def tier1_and_manual_still_outrank_the_plate():
    # The same tracker saw the whole movement: tier 1 claims it first and the plate
    # tier must not double-pair either end.
    write([("cam1", 1, "car", GE, "entry", ts("08:00")),
           ("cam1", 1, "car", GE, "exit", ts("08:00", 20)),
           ("cam2", 51, "car", CH, "exit", ts("08:00", 30))])
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4021")])
    d = run()
    assert d["movements"]["tier1"] == 1 and d["movements"]["tier3"] == 0, d["movements"]


def a_plate_pair_prices_as_inferred_per_cell():
    # Only the plate-bearing pair on the books: its cell must wear the conservative
    # "inferred" per-cell label (tier >= 2 fold), with the split told in Method/QA.
    write([("cam1", 1, "car", GE, "entry", ts("08:00")),
           ("cam2", 51, "car", CH, "exit", ts("08:00", 30))])
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "BAE4021"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "BAE4021")])
    d = run()
    assert d["movements"]["tier3"] == 1 and d["movements"]["tier2"] == 0, d["movements"]
    cell = [c for c in d["movements"]["od"] if c["count"]][0]
    assert cell["count"] == 1 and cell["tier2_count"] == 1, cell


def plate_strings_reach_no_payload_and_no_workbook():
    write(AMBIG)
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "ZDPA1234"),
            prow("cam2", 51, CH, "exit", ts("08:00", 30), "ZDPA1234")])
    d = run()
    assert "ZDPA1234" not in json.dumps(d), "plate leaked into the counts payload"
    for ep in (f"/api/refine/events/{SITE}/{DATE}",
               f"/api/pair/candidates/{SITE}/{DATE}?cam=cam1&obj_id=1"
               f"&line={GE.replace(' ', '%20')}&kind=entry&ts={ts('08:00')}"):
        r = C.get(ep)
        assert r.status_code == 200 and "ZDPA1234" not in r.text, ep
    out = report.build(SITE, DATE, app.data_root(), TMP / "ingest", app.CONFIG)
    xlsx = (Path(out["dir"]) / f"countkit-{SITE}-{DATE}.xlsx").read_bytes()
    assert b"ZDPA1234" not in xlsx, "plate leaked into the client workbook"
    # But the workbook DOES disclose that the tier exists.
    from openpyxl import load_workbook
    ws = load_workbook(Path(out["dir"]) / f"countkit-{SITE}-{DATE}.xlsx")["Method"]
    rows = {r[0]: r[1] for r in ws.iter_rows(values_only=True) if r and r[0]}
    assert rows.get("Plate-confirmed pairs") == 1, rows.get("Plate-confirmed pairs")


def reanalysis_wipes_machine_sidecars_but_not_human_ledgers():
    """Derived plate strings must not outlive their evidence custody; a reviewer's
    refine ledger must survive and resolve (or evaporate visibly) on its own."""
    plates([prow("cam1", 1, GE, "entry", ts("08:00"), "ZDPA9999")])
    refine = TMP / "data" / "refine" / f"{SITE}-{DATE}.jsonl"
    refine.parent.mkdir(parents=True, exist_ok=True)
    refine.write_text('{"cam": "cam1", "obj_id": 1, "to": "lgv"}\n')
    # The smallest runnable analysis: a verified site-day with offsets and a mock
    # detector holding zero segments — check_ready and detector construction pass,
    # so the idempotent cleanup actually runs.
    day = TMP / "ingest" / DATE / SITE
    (day / "cam1").mkdir(parents=True, exist_ok=True)
    (day / ".verified").touch()
    (day / "manifest.json").write_text(json.dumps(
        {"time_offset_s": {"cam1": 0.0, "cam2": 0.0}}))   # cam2: module harness made it
    fix = TMP / "fixtures" / SITE / DATE
    fix.mkdir(parents=True, exist_ok=True)
    (fix / "segments.json").write_text(json.dumps({"cam1": [], "cam2": []}))
    engine.analyze(SITE, DATE, TMP / "ingest", app.data_root(),
                   engine.mock_factory(TMP / "fixtures"))
    assert not (TMP / "data" / "plate" / f"{SITE}-{DATE}.jsonl").exists(), \
        "plate strings survived the analysis that orphaned them"
    assert refine.exists(), "a human ledger was destroyed by re-analysis"
    refine.write_text("")


def the_distance_helper_is_exact():
    ok, no = aggregate._plate_dist_ok, lambda a, b: not aggregate._plate_dist_ok(a, b)
    assert ok("BAE4021", "BAE4021") and ok("BAE4021", "BAE4027")
    assert ok("BAE4021", "BAE402") and ok("BAE402", "BAE4021") and ok("BAE4021", "AE4021")
    assert no("BAE4021", "BAF4027") and no("BAE4021", "BAE40") and no("ABC", "XYZ")
    assert no("ABAB", "BABA")       # transposition-heavy lookalikes stay apart


check("a plate settles what the gate refused", a_plate_settles_what_the_gate_refused)
check("one OCR slip is tolerated, two are not", one_ocr_slip_is_tolerated_two_are_not)
check("a plate in reach of two exits binds nothing",
      a_plate_in_reach_of_two_exits_binds_nothing)
check("a low-confidence read is no read", a_low_confidence_read_is_no_read)
check("plates outrank a one-sided refinement", plates_outrank_a_one_sided_refinement)
check("a stale sidecar shows reads without resolution",
      a_stale_sidecar_shows_reads_without_resolution)
check("tier 1 and manual still outrank the plate", tier1_and_manual_still_outrank_the_plate)
check("a plate pair prices as inferred per cell", a_plate_pair_prices_as_inferred_per_cell)
check("plate strings reach no payload and no workbook",
      plate_strings_reach_no_payload_and_no_workbook)
check("reanalysis wipes machine sidecars but not human ledgers",
      reanalysis_wipes_machine_sidecars_but_not_human_ledgers)
check("the distance helper is exact", the_distance_helper_is_exact)

print(f"\n{'FAILED: ' + ', '.join(FAILS) if FAILS else 'all passed'}")
sys.exit(1 if FAILS else 0)
